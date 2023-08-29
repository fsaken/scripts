import requests
import openpyxl
from openpyxl.styles import PatternFill,Font

API_EMAIL = "YOUR-EMAIL-LOGIN"
API_KEY = "YOUR-API-KEY"

API_BASE_URL = "https://api.cloudflare.com/client/v4"

def get_zones():
    headers = {
        "X-Auth-Email": API_EMAIL,
        "X-Auth-Key": API_KEY
    }

    response = requests.get(f"{API_BASE_URL}/zones", headers=headers)
    data = response.json()

    if response.status_code == 200 and data.get("success", False):
        zones = data.get("result", [])
        return zones
    else:
        print("Erro ao obter as zonas da API")
        print(data)
        return []

def get_waf_rules(zone_id):
    headers = {
        "X-Auth-Email": API_EMAIL,
        "X-Auth-Key": API_KEY
    }

    response = requests.get(f"{API_BASE_URL}/zones/{zone_id}/firewall/rules", headers=headers)
    data = response.json()

    if response.status_code == 200 and data.get("success", False):
        waf_rules = data.get("result", [])
        return waf_rules
    else:
        print("Erro ao obter as regras WAF da API")
        print(data)
        return []

if __name__ == "__main__":
    zones = get_zones()

    wb = openpyxl.Workbook()

    for zone in zones:
        zone_id = zone["id"]
        zone_name = zone["name"]
        
        ws = wb.create_sheet(title=zone_name)
        
        ws.append(["WAF Rule ID", "Description", "Action", "Expression"])
        
        table_header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        table_header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = table_header_fill
            cell.font = table_header_font
        
        waf_rules = get_waf_rules(zone_id)
        
        for waf_rule in waf_rules:
            ws.append([waf_rule['id'], waf_rule['description'],  waf_rule['action'],waf_rule['filter']['expression']])
    
    wb.remove(wb['Sheet'])
    
    for sheet in wb:
        for column_cells in sheet.columns:
            max_length = 0
            for cell in column_cells:
                try:
                    value = cell.value
                    if value is not None:
                        max_length = max(max_length, len(str(value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width
    
    wb.save("WAF Rules.xlsx")